import random
import numpy as np
import matplotlib.pyplot as plt
from typing import List, Tuple, Dict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO, StringIO
import sys

class AlgoritmoGenetico:
    
    def __init__(self, 
                 num_vigilantes: int = 4,
                 num_dias: int = 7,
                 num_turnos: int = 2,
                 num_puestos: int = 3,
                 tam_poblacion: int = 20,
                 prob_cruza: float = 0.85,
                 prob_mutacion_individuo: float = 0.20,
                 prob_mutacion_gen: float = 0.25,
                 max_generaciones: int = 100,
                 horas_max: float = 60.0,
                 max_turnos_nocturnos: int = 3,
                 datos_vigilantes: List[Dict] = None):
        
        self.num_vigilantes = num_vigilantes
        self.num_dias = num_dias
        self.num_turnos = num_turnos
        self.num_puestos = num_puestos
        self.tam_poblacion = tam_poblacion
        self.prob_cruza = prob_cruza
        self.prob_mutacion_individuo = prob_mutacion_individuo
        self.prob_mutacion_gen = prob_mutacion_gen
        self.max_generaciones = max_generaciones
        self.horas_max = horas_max
        self.max_turnos_nocturnos = max_turnos_nocturnos
        self.datos_vigilantes = datos_vigilantes or []
        
        self.log_buffer = StringIO()
        self.stdout_original = sys.stdout
        sys.stdout = self.log_buffer
        
        self.longitud_cromosoma = num_dias * num_turnos * num_puestos
        self.duraciones_turno = [10.5, 10.5]
        self.horarios_turno = ['08:00-20:00 (Desayuno 10:00-11:00, Comida 16:00-16:30)', '20:00-08:00 (Desayuno 22:00-23:00, Comida 02:00-02:30)']
        self.es_nocturno = [0, 1]
        self.nombres_puestos = ['Entrada', 'Salida', 'Estacionamiento']
        self.nombres_dias = ['Lunes', 'Martes', 'Miercoles', 'Jueves', 'Viernes', 'Sabado', 'Domingo']
        self.es_domingo = [0, 0, 0, 0, 0, 0, 1]
        self.requerimientos = [[3, 3, 3], [3, 3, 3], [3, 3, 3], [3, 3, 3], [3, 3, 3], [3, 3, 3], [3, 3, 3]]
        
        if self.datos_vigilantes:
            print(f"\nDatos de vigilantes cargados: {len(self.datos_vigilantes)} vigilantes")
            for v in self.datos_vigilantes[:3]:
                print(f"  - {v.get('Nombre', 'N/A')} {v.get('Apellido', 'N/A')} (Turno preferido: {v.get('Turno_Preferido', 'N/A')})")
            if len(self.datos_vigilantes) > 3:
                print(f"  ... y {len(self.datos_vigilantes) - 3} mas")
        
        print(f"Algoritmo Genetico Inicializado")
    
    def obtener_nombre_vigilante(self, id_vigilante: int) -> str:
        if self.datos_vigilantes and 0 < id_vigilante <= len(self.datos_vigilantes):
            v = self.datos_vigilantes[id_vigilante - 1]
            nombre = v.get('Nombre', f'Vigilante {id_vigilante}')
            apellido = v.get('Apellido', '')
            return f"{nombre} {apellido}".strip()
        return f"Vigilante {id_vigilante}"
    
    def obtener_info_vigilante(self, id_vigilante: int) -> Dict:
        if self.datos_vigilantes and 0 < id_vigilante <= len(self.datos_vigilantes):
            return self.datos_vigilantes[id_vigilante - 1]
        return {}
        print(f"Longitud cromosoma: {self.longitud_cromosoma} genes")
        print(f"Poblacion: {tam_poblacion} individuos")
        print(f"Vigilantes: {num_vigilantes}")
        print(f"Configuracion: {num_dias} días, {num_turnos} turnos, {num_puestos} puestos")

    def generar_individuo(self) -> List[int]:
        return [random.randint(1, self.num_vigilantes) 
                for _ in range(self.longitud_cromosoma)]
    
    def inicializar_poblacion(self) -> List[List[int]]:
        poblacion = [self.generar_individuo() 
                     for _ in range(self.tam_poblacion)]
        print(f"Poblacion inicial generada: {len(poblacion)} individuos")
        return poblacion

    def calcular_cobertura_turnos(self, cromosoma: List[int]) -> float:
        turnos_requeridos_total = 0
        turnos_cubiertos_total = 0
        
        for dia in range(self.num_dias):
            for turno in range(self.num_turnos):
                puestos_requeridos = self.requerimientos[dia][turno] if dia < len(self.requerimientos) and turno < len(self.requerimientos[dia]) else self.num_puestos
                turnos_requeridos_total += puestos_requeridos
                
                puestos_cubiertos = 0
                for puesto in range(self.num_puestos):
                    idx = dia * (self.num_turnos * self.num_puestos) + turno * self.num_puestos + puesto
                    if idx < len(cromosoma) and cromosoma[idx] > 0:
                        puestos_cubiertos += 1
                
                turnos_cubiertos_total += min(puestos_cubiertos, puestos_requeridos)
        
        CT = turnos_cubiertos_total / turnos_requeridos_total if turnos_requeridos_total > 0 else 1.0
        return min(CT, 1.0)
    
    def calcular_horas_trabajadas(self, cromosoma: List[int]) -> List[float]:
        horas = [0.0] * (self.num_vigilantes + 1)
        
        for dia in range(self.num_dias):
            for turno in range(self.num_turnos):
                for puesto in range(self.num_puestos):
                    idx = dia * (self.num_turnos * self.num_puestos) + \
                          turno * self.num_puestos + puesto
                    vigilante = cromosoma[idx]
                    duracion = self.duraciones_turno[turno]
                    horas[vigilante] += duracion
        
        return horas[1:]
    
    def calcular_horas_extra(self, cromosoma: List[int]) -> float:
        horas_extra_totales = 0.0
        horas_trabajadas = self.calcular_horas_trabajadas(cromosoma)
        
        horas_legales_semanales = 48.0
        
        for horas_vigilante in horas_trabajadas:
            if horas_vigilante > horas_legales_semanales:
                horas_extra_totales += (horas_vigilante - horas_legales_semanales)
        
        return horas_extra_totales
    
    def calcular_balance_carga_laboral(self, cromosoma: List[int]) -> float:
        horas = self.calcular_horas_trabajadas(cromosoma)
        if len(horas) == 0:
            return 0.0
        BCL = max(horas) - min(horas)
        return BCL
    
    def calcular_turnos_nocturnos(self, cromosoma: List[int]) -> List[int]:
        nocturnos = [0] * (self.num_vigilantes + 1)
        
        for dia in range(self.num_dias):
            for turno in range(self.num_turnos):
                if self.es_nocturno[turno] == 1:
                    for puesto in range(self.num_puestos):
                        idx = dia * (self.num_turnos * self.num_puestos) + \
                              turno * self.num_puestos + puesto
                        vigilante = cromosoma[idx]
                        nocturnos[vigilante] += 1
        
        return nocturnos[1:]
    
    def calcular_violaciones(self, cromosoma: List[int]) -> Tuple[int, bool]:
        violaciones = 0
        violaciones_graves = False
        
        for dia in range(self.num_dias):
            for vigilante in range(1, self.num_vigilantes + 1):
                turnos_trabajados = 0
                for turno in range(self.num_turnos):
                    for puesto in range(self.num_puestos):
                        idx = dia * (self.num_turnos * self.num_puestos) + \
                              turno * self.num_puestos + puesto
                        if cromosoma[idx] == vigilante:
                            turnos_trabajados += 1
                
                if turnos_trabajados > 2:
                    violaciones += (turnos_trabajados - 2) * 2
                    violaciones_graves = True
        
        nocturnos = self.calcular_turnos_nocturnos(cromosoma)
        for n in nocturnos:
            if n > self.max_turnos_nocturnos:
                violaciones += (n - self.max_turnos_nocturnos)
        
        return violaciones, violaciones_graves
    
    def funcion_aptitud(self, cromosoma: List[int]) -> float:
        CT = self.calcular_cobertura_turnos(cromosoma)
        HE = self.calcular_horas_extra(cromosoma)
        BCL = self.calcular_balance_carga_laboral(cromosoma)
        VL, violaciones_graves = self.calcular_violaciones(cromosoma)
        
        max_horas_extra_posibles = self.num_dias * self.num_vigilantes * 12
        max_balance = self.horas_max * 2
        max_violaciones = self.num_dias * self.num_vigilantes * 2
        
        penalizacion_HE = min(1.0, HE / max_horas_extra_posibles) if max_horas_extra_posibles > 0 else 0
        penalizacion_BCL = min(1.0, BCL / max_balance) if max_balance > 0 else 0
        penalizacion_VL = min(1.0, VL / max_violaciones) if max_violaciones > 0 else 0
        
        fitness = CT * (1 - 0.3 * penalizacion_HE - 0.2 * penalizacion_BCL - 0.5 * penalizacion_VL)
        
        fitness = max(0.0001, min(1.0, fitness))
        
        return fitness

    def seleccionar_mejores(self, poblacion: List[List[int]], 
                           aptitudes: List[float], n: int = 3) -> List[int]:
        indices_ordenados = sorted(range(len(aptitudes)), 
                                  key=lambda i: aptitudes[i], 
                                  reverse=True)
        return indices_ordenados[:n]
    
    def seleccionar_peor(self, poblacion: List[List[int]], 
                        aptitudes: List[float]) -> int:
        return aptitudes.index(min(aptitudes))

    def generar_parejas(self, poblacion: List[List[int]], 
                       aptitudes: List[float]) -> List[Tuple[int, int]]:
        n = len(poblacion)
        parejas = []
        
        todas_parejas = []
        for i in range(n):
            for j in range(i + 1, n):
                todas_parejas.append((i, j))
        
        random.shuffle(todas_parejas)
        
        usados = set()
        for i, j in todas_parejas:
            if i not in usados and j not in usados:
                parejas.append((i, j))
                usados.add(i)
                usados.add(j)
        
        return parejas

    def cruza_un_punto(self, padre1: List[int], padre2: List[int]) -> Tuple[List[int], List[int]]:
        punto_cruza = random.randint(1, len(padre1) - 1)
        hijo1 = padre1[:punto_cruza] + padre2[punto_cruza:]
        hijo2 = padre2[:punto_cruza] + padre1[punto_cruza:]
        return hijo1, hijo2
    
    def aplicar_cruza(self, poblacion: List[List[int]], 
                     parejas: List[Tuple[int, int]]) -> List[List[int]]:
        hijos = []
        
        for i, j in parejas:
            p = random.random()
            
            if p <= self.prob_cruza:
                hijo1, hijo2 = self.cruza_un_punto(poblacion[i], poblacion[j])
                hijos.extend([hijo1, hijo2])
            else:
                hijos.extend([poblacion[i].copy(), poblacion[j].copy()])
        
        return hijos

    def mutar_individuo(self, individuo: List[int]) -> List[int]:
        mutado = individuo.copy()
        
        for i in range(len(mutado)):
            if random.random() <= self.prob_mutacion_gen:
                nuevo_valor = random.randint(1, self.num_vigilantes)
                mutado[i] = nuevo_valor
        
        return mutado
    
    def aplicar_mutacion(self, poblacion: List[List[int]]) -> List[List[int]]:
        poblacion_mutada = []
        
        for individuo in poblacion:
            if random.random() <= self.prob_mutacion_individuo:
                poblacion_mutada.append(self.mutar_individuo(individuo))
            else:
                poblacion_mutada.append(individuo.copy())
        
        return poblacion_mutada

    def aplicar_poda(self, poblacion: List[List[int]], 
                    aptitudes: List[float]) -> Tuple[List[List[int]], List[float]]:
        
        while len(poblacion) > self.tam_poblacion:
            peor_idx = self.seleccionar_peor(poblacion, aptitudes)
            poblacion.pop(peor_idx)
            aptitudes.pop(peor_idx)
        
        return poblacion, aptitudes

    def mostrar_grafica(self, historial: List[float]):
        plt.figure(figsize=(12, 6))
        plt.plot(historial, linewidth=2, color='#2E86AB')
        plt.xlabel('Generacion', fontsize=12)
        plt.ylabel('Fitness', fontsize=12)
        plt.title('Evolucion del Fitness', fontsize=14, fontweight='bold')
        plt.grid(True, alpha=0.3)
        plt.ylim(0, 1.05)
        
        mejor_fitness = max(historial)
        mejor_gen = historial.index(mejor_fitness)
        plt.axhline(y=mejor_fitness, color='r', linestyle='--', alpha=0.5, label=f'Mejor: {mejor_fitness:.4f}')
        plt.plot(mejor_gen, mejor_fitness, 'ro', markersize=10, label=f'Gen {mejor_gen}')
        
        plt.legend()
        plt.tight_layout()
        plt.show()

    def exportar_excel(self, poblacion_final: List[List[int]], aptitudes_finales: List[float], 
                      historial: List[float], nombre_archivo: str = None):
        if nombre_archivo is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nombre_archivo = f"resultados_ag_{timestamp}.xlsx"
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # Colores y estilos
        fill_titulo = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
        fill_header = PatternFill(start_color="4A90A4", end_color="4A90A4", fill_type="solid")
        fill_dia = PatternFill(start_color="D4E6F1", end_color="D4E6F1", fill_type="solid")
        fill_mejor = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fill_peor = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        font_titulo = Font(bold=True, size=14, color="FFFFFF")
        font_header = Font(bold=True, size=11, color="FFFFFF")
        font_bold = Font(bold=True)
        
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Obtener mejores soluciones
        indices_ordenados = sorted(range(len(aptitudes_finales)), 
                                  key=lambda i: aptitudes_finales[i], 
                                  reverse=True)
        mejores_3 = indices_ordenados[:3]
        
        # Crear sheet de resumen
        ws_resumen = wb.create_sheet('Resumen Comparativo', 0)
        ws_resumen.column_dimensions['A'].width = 15
        
        # Título
        ws_resumen.merge_cells('A1:F1')
        cell = ws_resumen['A1']
        cell.value = 'COMPARACION DE SOLUCIONES'
        cell.font = font_titulo
        cell.fill = fill_titulo
        cell.alignment = center
        ws_resumen.row_dimensions[1].height = 25
        
        # Encabezados
        headers = ['Ranking', 'Fitness', 'Cobertura', 'Horas Extra', 'Balance', 'Violaciones']
        for col, header in enumerate(headers, 1):
            cell = ws_resumen.cell(row=3, column=col)
            cell.value = header
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = center
            cell.border = thin_border
        
        # Datos Top 1, 2, 3
        for rank_pos, idx in enumerate(mejores_3):
            cromosoma = poblacion_final[idx]
            fitness = aptitudes_finales[idx]
            CT = self.calcular_cobertura_turnos(cromosoma)
            HE = self.calcular_horas_extra(cromosoma)
            BCL = self.calcular_balance_carga_laboral(cromosoma)
            VL, _ = self.calcular_violaciones(cromosoma)
            
            row = 4 + rank_pos
            cells_data = [f'Top {rank_pos + 1}', fitness, CT, HE, BCL, VL]
            
            for col, value in enumerate(cells_data, 1):
                cell = ws_resumen.cell(row=row, column=col)
                cell.value = value
                cell.fill = fill_mejor
                cell.font = font_bold
                cell.alignment = center
                cell.border = thin_border
                if isinstance(value, float):
                    cell.number_format = '0.0000'
        
        # Para cada solución Top 1, 2, 3, crear sheet de calendario
        for rank_idx, idx in enumerate(mejores_3):
            ws = wb.create_sheet(f'Top {rank_idx + 1} - Calendario')
            
            cromosoma = poblacion_final[idx]
            fitness = aptitudes_finales[idx]
            CT = self.calcular_cobertura_turnos(cromosoma)
            HE = self.calcular_horas_extra(cromosoma)
            BCL = self.calcular_balance_carga_laboral(cromosoma)
            VL, _ = self.calcular_violaciones(cromosoma)
            
            # Título
            ws.merge_cells('A1:F1')
            cell = ws['A1']
            cell.value = f'HORARIO TOP {rank_idx + 1} - FITNESS: {fitness:.4f}'
            cell.font = font_titulo
            cell.fill = fill_titulo
            cell.alignment = center
            ws.row_dimensions[1].height = 25
            
            # Métricas
            ws['A2'] = 'Cobertura:'
            ws['B2'] = f'{CT:.2%}'
            ws['C2'] = 'Horas Extra:'
            ws['D2'] = f'{HE:.1f}h'
            ws['E2'] = 'Balance:'
            ws['F2'] = f'{BCL:.1f}h'
            
            fila = 4
            ws.merge_cells(f'A{fila}:F{fila}')
            cell = ws[f'A{fila}']
            cell.value = 'CALENDARIO DE TURNOS'
            cell.font = font_titulo
            cell.fill = fill_titulo
            cell.alignment = center
            fila += 2
            
            cromosoma_idx = 0
            for dia in range(self.num_dias):
                # Nombre del día
                ws.merge_cells(f'A{fila}:F{fila}')
                cell = ws[f'A{fila}']
                cell.value = f'DIA {dia + 1}: {self.nombres_dias[dia]}'
                cell.font = font_bold
                cell.fill = fill_dia
                cell.alignment = center
                fila += 1
                
                for turno in range(self.num_turnos):
                    # Horario
                    hora_inicio = "08:00" if turno == 0 else "20:00"
                    hora_fin = "20:00" if turno == 0 else "08:00"
                    turno_nombre = f'Turno Día ({hora_inicio}-{hora_fin})'  if turno == 0 else f'Turno Noche ({hora_inicio}-{hora_fin})'
                    duracion = self.duraciones_turno[turno]
                    
                    ws.merge_cells(f'A{fila}:F{fila}')
                    cell = ws[f'A{fila}']
                    cell.value = f'{turno_nombre} - {duracion}h'
                    cell.font = font_bold
                    cell.alignment = center
                    fila += 1
                    
                    # Encabezados de puesto
                    headers_puesto = ['Puesto', 'Vigilante', 'Horas', 'Tipo', 'Estado']
                    for col, header in enumerate(headers_puesto, 1):
                        cell = ws.cell(row=fila, column=col)
                        cell.value = header
                        cell.font = font_header
                        cell.fill = fill_header
                        cell.alignment = center
                        cell.border = thin_border
                    fila += 1
                    
                    # Puestos
                    for puesto in range(self.num_puestos):
                        vigilante = cromosoma[cromosoma_idx]
                        cromosoma_idx += 1
                        
                        ws.cell(row=fila, column=1).value = self.nombres_puestos[puesto]
                        ws.cell(row=fila, column=2).value = f'Vigilante {vigilante}'
                        ws.cell(row=fila, column=3).value = duracion
                        ws.cell(row=fila, column=4).value = 'Nocturno' if turno == 1 else 'Diurno'
                        
                        estado = "Normal"
                        if self.es_domingo[dia]:
                            estado = "Extra (Domingo)"
                        
                        ws.cell(row=fila, column=5).value = estado
                        
                        for col in range(1, 6):
                            ws.cell(row=fila, column=col).alignment = center
                            ws.cell(row=fila, column=col).border = thin_border
                        
                        fila += 1
                    
                    fila += 1
            
            fila += 2
            ws.merge_cells(f'A{fila}:F{fila}')
            cell = ws[f'A{fila}']
            cell.value = 'CROMOSOMA (GENES)'
            cell.font = font_titulo
            cell.fill = fill_titulo
            cell.alignment = center
            fila += 1
            
            headers_genes = ['Índice', 'Día', 'Turno', 'Puesto', 'Vigilante ID', 'Valor Gen']
            for col, header in enumerate(headers_genes, 1):
                cell = ws.cell(row=fila, column=col)
                cell.value = header
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = center
                cell.border = thin_border
            fila += 1
            
            cromosoma_idx = 0
            for dia in range(self.num_dias):
                for turno in range(self.num_turnos):
                    for puesto in range(self.num_puestos):
                        gen_value = cromosoma[cromosoma_idx]
                        ws.cell(row=fila, column=1).value = cromosoma_idx
                        ws.cell(row=fila, column=2).value = self.nombres_dias[dia]
                        ws.cell(row=fila, column=3).value = 'Noche' if turno == 1 else 'Día'
                        ws.cell(row=fila, column=4).value = self.nombres_puestos[puesto]
                        ws.cell(row=fila, column=5).value = gen_value
                        ws.cell(row=fila, column=6).value = gen_value
                        
                        for col in range(1, 7):
                            ws.cell(row=fila, column=col).alignment = center
                            ws.cell(row=fila, column=col).border = thin_border
                        
                        cromosoma_idx += 1
                        fila += 1
        
        ws_consola = wb.create_sheet('Consola', len(wb.sheetnames))
        
        fill_titulo_cons = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        font_titulo_cons = Font(bold=True, size=12, color="FFFFFF")
        
        ws_consola['A1'] = 'SALIDA DE CONSOLA'
        ws_consola['A1'].font = font_titulo_cons
        ws_consola['A1'].fill = fill_titulo_cons
        ws_consola['A1'].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws_consola.column_dimensions['A'].width = 120
        
        logs = self.log_buffer.getvalue()
        log_lines = logs.split('\n')
        
        fila_log = 3
        for line in log_lines:
            if line.strip():
                cell = ws_consola[f'A{fila_log}']
                cell.value = line
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.border = Border(bottom=Side(style='thin'))
                ws_consola.row_dimensions[fila_log].height = None
                fila_log += 1
        
        sys.stdout = self.stdout_original
        
        wb.save(nombre_archivo)
        print(f"\nArchivo Excel generado: {nombre_archivo}")
        return nombre_archivo

    def guardar_grafica(self, historial: List[float], nombre_archivo: str = None):
        if nombre_archivo is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nombre_archivo = f"evolucion_fitness_{timestamp}.png"
        
        plt.figure(figsize=(12, 6))
        plt.plot(historial, linewidth=2, color='#2E86AB')
        plt.xlabel('Generacion', fontsize=12)
        plt.ylabel('Fitness', fontsize=12)
        plt.title('Evolucion del Fitness', fontsize=14, fontweight='bold')
        plt.grid(True, alpha=0.3)
        plt.ylim(0, 1.05)
        
        mejor_fitness = max(historial)
        mejor_gen = historial.index(mejor_fitness)
        plt.axhline(y=mejor_fitness, color='r', linestyle='--', alpha=0.5, label=f'Mejor: {mejor_fitness:.4f}')
        plt.plot(mejor_gen, mejor_fitness, 'ro', markersize=10, label=f'Gen {mejor_gen}')
        
        plt.legend()
        plt.tight_layout()
        plt.savefig(nombre_archivo, dpi=300, bbox_inches='tight')
        print(f"\nGrafica guardada: {nombre_archivo}")
        plt.close()

    def ejecutar(self) -> Tuple[List[int], float, List[int], float, List[int], float, List[int], float, List[float], List[List[int]], List[float]]:
        print("\n" + "="*60)
        print("INICIANDO ALGORITMO GENETICO")
        print("="*60 + "\n")
        
        poblacion = self.inicializar_poblacion()
        historial_aptitudes = []
        
        # Inicializar los 4 mejores globales
        top_1_global = self.generar_individuo()
        top_1_fitness = 0.0
        top_2_global = self.generar_individuo()
        top_2_fitness = 0.0
        top_3_global = self.generar_individuo()
        top_3_fitness = 0.0
        peor_global = self.generar_individuo()
        peor_fitness = float('inf')
        
        for generacion in range(self.max_generaciones):
            print(f"\nGeneracion {generacion + 1}/{self.max_generaciones}")
            
            aptitudes = [self.funcion_aptitud(ind) for ind in poblacion]
            
            mejor_idx = aptitudes.index(max(aptitudes))
            mejor_aptitud = aptitudes[mejor_idx]
            
            # Actualizar Top 1 (mejor global)
            if mejor_aptitud > top_1_fitness:
                top_1_fitness = mejor_aptitud
                top_1_global = poblacion[mejor_idx].copy()
                print(f"  Nuevo Top 1 global: {top_1_fitness:.4f}")
            
            # Actualizar Top 2 (segundo mejor global)
            if mejor_aptitud > top_2_fitness and mejor_aptitud < top_1_fitness:
                top_2_fitness = mejor_aptitud
                top_2_global = poblacion[mejor_idx].copy()
                print(f"  Nuevo Top 2 global: {top_2_fitness:.4f}")
            elif mejor_aptitud > top_2_fitness and mejor_aptitud == top_1_fitness:
                # Si es igual al top 1, buscar el segundo mejor de esta generacion
                aptitudes_ordenadas = sorted(enumerate(aptitudes), key=lambda x: x[1], reverse=True)
                if len(aptitudes_ordenadas) > 1:
                    segundo_idx = aptitudes_ordenadas[1][0]
                    segundo_fitness = aptitudes_ordenadas[1][1]
                    if segundo_fitness > top_2_fitness:
                        top_2_fitness = segundo_fitness
                        top_2_global = poblacion[segundo_idx].copy()
                        print(f"  Nuevo Top 2 global: {top_2_fitness:.4f}")
            
            # Actualizar Top 3 (tercer mejor global)
            aptitudes_ordenadas = sorted(enumerate(aptitudes), key=lambda x: x[1], reverse=True)
            if len(aptitudes_ordenadas) >= 3:
                for idx, fitness in aptitudes_ordenadas:
                    if fitness > top_3_fitness and fitness < top_2_fitness:
                        top_3_fitness = fitness
                        top_3_global = poblacion[idx].copy()
                        print(f"  Nuevo Top 3 global: {top_3_fitness:.4f}")
                        break
            elif len(aptitudes_ordenadas) >= 2:
                # Si solo hay 2 individuos, el tercero es el segundo
                segundo_idx = aptitudes_ordenadas[1][0]
                segundo_fitness = aptitudes_ordenadas[1][1]
                if segundo_fitness > top_3_fitness:
                    top_3_fitness = segundo_fitness
                    top_3_global = poblacion[segundo_idx].copy()
            
            # Actualizar Peor (peor global - el de menor fitness encontrado)
            peor_idx = aptitudes.index(min(aptitudes))
            peor_aptitud = aptitudes[peor_idx]
            if peor_aptitud < peor_fitness:
                peor_fitness = peor_aptitud
                peor_global = poblacion[peor_idx].copy()
                print(f"  Nuevo Peor global: {peor_fitness:.4f}")
            
            # CORRECCION: Guardar el mejor de ESTA generacion, no el global acumulativo
            historial_aptitudes.append(mejor_aptitud)
            
            print(f"  Mejor aptitud: {mejor_aptitud:.4f}")
            print(f"  Aptitud promedio: {np.mean(aptitudes):.4f}")
            
            mejores_indices = self.seleccionar_mejores(poblacion, aptitudes, 3)
            print(f"  Top 3 aptitudes: {[f'{aptitudes[i]:.4f}' for i in mejores_indices]}")
            
            parejas = self.generar_parejas(poblacion, aptitudes)
            hijos = self.aplicar_cruza(poblacion, parejas)
            poblacion = self.aplicar_mutacion(hijos)
            
            aptitudes = [self.funcion_aptitud(ind) for ind in poblacion]
            poblacion, aptitudes = self.aplicar_poda(poblacion, aptitudes)
        
        print("\n" + "="*60)
        print("ALGORITMO FINALIZADO")
        print("="*60)
        print(f"Top 1 Global: {top_1_fitness:.4f}")
        print(f"Top 2 Global: {top_2_fitness:.4f}")
        print(f"Top 3 Global: {top_3_fitness:.4f}")
        print(f"Peor Global: {peor_fitness:.4f}")
        
        self.poblacion_final = poblacion
        self.aptitudes_finales = aptitudes
        self.historial = historial_aptitudes
        
        return (top_1_global, top_1_fitness, 
                top_2_global, top_2_fitness, 
                top_3_global, top_3_fitness, 
                peor_global, peor_fitness, 
                historial_aptitudes, poblacion, aptitudes)
    
    def imprimir_solucion(self, cromosoma: List[int]):
        print("\nSOLUCION DETALLADA:")
        print("-" * 60)
        
        idx = 0
        for dia in range(self.num_dias):
            print(f"\n{self.nombres_dias[dia].upper()} (DIA {dia + 1})")
            for turno in range(self.num_turnos):
                turno_nombre = "Dia (08:00-20:00)" if turno == 0 else "Noche (20:00-08:00)"
                print(f"  Turno: {turno_nombre}")
                if turno == 0:
                    print(f"    Descansos: Desayuno 10:00-11:00, Comida 16:00-16:30")
                else:
                    print(f"    Descansos: Desayuno 22:00-23:00, Comida 02:00-02:30")
                print(f"    Horas efectivas: 10.5h")
                for puesto in range(self.num_puestos):
                    vigilante = cromosoma[idx]
                    print(f"    {self.nombres_puestos[puesto]}: Vigilante {vigilante}")
                    idx += 1
        
        CT = self.calcular_cobertura_turnos(cromosoma)
        HE = self.calcular_horas_extra(cromosoma)
        BCL = self.calcular_balance_carga_laboral(cromosoma)
        VL, violaciones_graves = self.calcular_violaciones(cromosoma)
        horas = self.calcular_horas_trabajadas(cromosoma)
        nocturnos = self.calcular_turnos_nocturnos(cromosoma)
        fitness = self.funcion_aptitud(cromosoma)
        
        print("\nMETRICAS:")
        print("-" * 60)
        print(f"FITNESS TOTAL: {fitness:.4f}")
        print(f"Cobertura (CT): {CT:.2%}")
        print(f"Horas Extra (HE): {HE:.1f} horas")
        print(f"Balance (BCL): {BCL:.1f} horas")
        print(f"Violaciones (VL): {VL}")
        
        if VL > 0:
            print(f"\nNOTA: Hay {VL} violaciones")
            print(f"      Permitido: hasta 2 turnos/dia por vigilante")
            print(f"      Mas de 2 turnos/dia se penaliza fuertemente")
        
        print("\nHORAS TRABAJADAS:")
        horas_legales_semanales = 48.0
        for i, h in enumerate(horas, 1):
            print(f"  Vigilante {i}: {h:.1f}h total", end="")
            
            horas_extra_vigilante = max(0, h - horas_legales_semanales)
            
            if horas_extra_vigilante > 0:
                print(f" (Extra: {horas_extra_vigilante:.1f}h sobre 48h legales)", end="")
            print()
        
        print("\nTURNOS NOCTURNOS:")
        for i, n in enumerate(nocturnos, 1):
            estado = "OK" if n <= self.max_turnos_nocturnos else "EXCESO"
            print(f"  Vigilante {i}: {n} turnos [{estado}]")


if __name__ == "__main__":
    ag = AlgoritmoGenetico(
        num_vigilantes=4,
        num_dias=7,
        num_turnos=2,
        num_puestos=3,
        tam_poblacion=30,
        prob_cruza=0.85,
        prob_mutacion_individuo=0.20,
        prob_mutacion_gen=0.15,
        max_generaciones=100,
        horas_max=60.0,
        max_turnos_nocturnos=7
    )
    
    mejor_solucion, mejor_aptitud, historial, poblacion_final, aptitudes_finales = ag.ejecutar()
    
    print(f"\nMEJOR APTITUD: {mejor_aptitud:.4f}")
    print(f"MEJOR CROMOSOMA: {mejor_solucion}")
    
    ag.imprimir_solucion(mejor_solucion)
    
    print("\nEVOLUCION DEL FITNESS:")
    print("-"*60)
    print(f"{'Generacion':<15} {'Fitness':<15} {'Estado':<20}")
    print("-"*60)
    for i in range(0, len(historial), 10):
        fitness = historial[i]
        if fitness >= 0.8:
            estado = "Excelente"
        elif fitness >= 0.5:
            estado = "Buena"
        elif fitness > 0:
            estado = "Mejorable"
        else:
            estado = "Sin solucion valida"
        print(f"Gen {i:<11} {fitness:<15.4f} {estado:<20}")
    
    ag.exportar_excel(poblacion_final, aptitudes_finales, historial)
    print("\nProceso completado. Presiona Enter para cerrar...")
    input()
